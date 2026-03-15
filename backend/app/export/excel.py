"""
Excel financial model export for KEEN due diligence reports.

Generates a structured .xlsx workbook from pipeline deliverables with:
  - Cover sheet (engagement metadata)
  - Executive Summary tab
  - Key Findings tab (color-coded by severity)
  - One data tab per source (Salesforce, NetSuite, Bloomberg, etc.)
  - Valuation Multiples tab
  - Compliance tab

Uses openpyxl for pure-Python Excel generation (no libreoffice required).
"""

from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side,
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Theme colours ─────────────────────────────────────────────────────────────
NAVY    = "0A1628"
BLUE    = "1E3A5F"
ACCENT  = "2563EB"
WHITE   = "FFFFFF"
LIGHT   = "F0F4F8"
BORDER  = "CBD5E1"
GREEN   = "16A34A"
AMBER   = "D97706"
RED     = "DC2626"
GREY    = "6B7280"

SEV_FILL = {
    "critical": PatternFill("solid", fgColor=RED),
    "warning":  PatternFill("solid", fgColor=AMBER),
    "info":     PatternFill("solid", fgColor="2563EB"),
}

# ── Helpers ───────────────────────────────────────────────────────────────────

def _header_font(size: int = 11, bold: bool = True, color: str = WHITE) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _body_font(size: int = 10, bold: bool = False, color: str = "1E293B") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)

def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _row(ws, values: list[Any], row: int, styles: dict | None = None) -> None:
    styles = styles or {}
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        if "font" in styles:      cell.font      = styles["font"]
        if "fill" in styles:      cell.fill      = styles["fill"]
        if "align" in styles:     cell.alignment = styles["align"]
        if "border" in styles:    cell.border    = styles["border"]
        if "number" in styles and isinstance(val, (int, float)):
            cell.number_format = styles["number"]

def _section_header(ws, title: str, row: int, ncols: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title.upper())
    c.font      = Font(name="Calibri", size=9, bold=True, color=GREY)
    c.alignment = _left()
    c.fill      = _fill(LIGHT)
    ws.row_dimensions[row].height = 18


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_cover(wb: Workbook, target_company: str, pe_firm: str, deliverables: dict) -> None:
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [4, 45, 35, 20])
    ws.row_dimensions[1].height = 10

    # Title block
    ws.merge_cells("B3:D3")
    c = ws.cell(row=3, column=2, value="KEEN DUE DILIGENCE PLATFORM")
    c.font = Font(name="Calibri", size=10, bold=True, color=GREY)

    ws.merge_cells("B4:D6")
    c = ws.cell(row=4, column=2, value=f"Financial Due Diligence\n{target_company}")
    c.font      = Font(name="Calibri", size=26, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 70

    ws.merge_cells("B8:D8")
    c = ws.cell(row=8, column=2, value=f"Prepared for {pe_firm}")
    c.font = Font(name="Calibri", size=13, color=BLUE)

    ws.merge_cells("B9:D9")
    c = ws.cell(row=9, column=2, value=datetime.now(timezone.utc).strftime("%B %d, %Y"))
    c.font = Font(name="Calibri", size=11, color=GREY)

    # Divider
    for col in range(2, 5):
        cell = ws.cell(row=11, column=col)
        cell.fill   = _fill(ACCENT)

    # Metadata table
    exec_summary = deliverables.get("executive_summary", {})
    meta_rows = [
        ("Engagement Type",   "Financial Due Diligence"),
        ("Target Company",    target_company),
        ("Client",            pe_firm),
        ("Report Date",       datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        ("Recommendation",    exec_summary.get("recommendation", "—")),
        ("Confidence Level",  exec_summary.get("confidence_level", "—")),
    ]

    ws.cell(row=13, column=2, value="ENGAGEMENT DETAILS").font = Font(name="Calibri", size=9, bold=True, color=GREY)
    for i, (label, value) in enumerate(meta_rows, start=14):
        ws.cell(row=i, column=2, value=label).font  = _body_font(bold=True)
        ws.cell(row=i, column=3, value=value).font  = _body_font()
        ws.row_dimensions[i].height = 18

    # Confidential notice
    ws.merge_cells("B22:D22")
    c = ws.cell(row=22, column=2,
                value="CONFIDENTIAL — FOR INTERNAL USE ONLY. "
                      "Not for distribution without prior written consent.")
    c.font      = Font(name="Calibri", size=9, italic=True, color=GREY)
    c.alignment = _left()


def _build_executive_summary(wb: Workbook, exec_summary: dict) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [3, 30, 55, 20])

    # Title
    ws.merge_cells("B2:D2")
    c = ws.cell(row=2, column=2, value="EXECUTIVE SUMMARY")
    c.font      = Font(name="Calibri", size=16, bold=True, color=NAVY)
    c.alignment = _left()

    # Recommendation banner
    rec = exec_summary.get("recommendation", "PROCEED")
    rec_color = {"PROCEED": GREEN, "PROCEED WITH CONDITIONS": AMBER, "DO NOT PROCEED": RED}.get(rec, ACCENT)
    ws.merge_cells("B4:D4")
    c = ws.cell(row=4, column=2, value=f"Recommendation: {rec}")
    c.font      = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill      = _fill(rec_color)
    c.alignment = _center()
    ws.row_dimensions[4].height = 28

    # KPI summary
    row = 6
    _section_header(ws, "KEY METRICS", row, 3)
    row += 1
    kpis = [
        ("Critical Findings",  exec_summary.get("critical_findings_count", 0)),
        ("Warning Findings",   exec_summary.get("warning_findings_count", 0)),
        ("Data Sources",       exec_summary.get("sources_analyzed", 0)),
        ("Confidence Level",   exec_summary.get("confidence_level", "—")),
    ]
    for label, value in kpis:
        ws.cell(row=row, column=2, value=label).font  = _body_font(bold=True)
        c = ws.cell(row=row, column=3, value=value)
        c.font = _body_font()
        c.border = _thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

    # Key findings narrative
    row += 1
    _section_header(ws, "KEY FINDINGS", row, 3)
    row += 1
    for finding in exec_summary.get("key_findings", []):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        c = ws.cell(row=row, column=2, value=f"• {finding}")
        c.font      = _body_font()
        c.alignment = _left()
        ws.row_dimensions[row].height = 22
        row += 1

    # Narrative paragraph
    row += 1
    _section_header(ws, "SUMMARY NARRATIVE", row, 3)
    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row + 3, end_column=3)
    c = ws.cell(row=row, column=2, value=exec_summary.get("narrative", ""))
    c.font      = _body_font()
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row].height = 80


def _build_findings(wb: Workbook, findings: list[dict]) -> None:
    ws = wb.create_sheet("Key Findings")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [3, 18, 12, 12, 45, 10])

    headers = ["Source System", "Type", "Severity", "Finding", "Review"]
    hrow = 2
    ws.cell(hrow, 2, "KEY FINDINGS").font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    hrow = 4
    for i, h in enumerate(headers, 2):
        c = ws.cell(row=hrow, column=i, value=h)
        c.font      = _header_font(color=WHITE)
        c.fill      = _fill(NAVY)
        c.alignment = _center()
        c.border    = _thin_border()
    ws.row_dimensions[hrow].height = 22

    for row_idx, finding in enumerate(findings, start=5):
        sev = finding.get("severity", "info")
        fill = SEV_FILL.get(sev, _fill(LIGHT))

        vals = [
            finding.get("source_system", ""),
            finding.get("type", finding.get("finding_type", "")),
            sev.upper(),
            finding.get("title", ""),
            "✓" if finding.get("requires_human_review") else "",
        ]
        for col, val in zip(range(2, 7), vals):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font      = _body_font()
            c.border    = _thin_border()
            c.alignment = _left()
            if col == 4:  # Severity column
                c.fill = fill
                c.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        ws.row_dimensions[row_idx].height = 20


def _build_data_tab(wb: Workbook, source_name: str, data: dict) -> None:
    """Create one tab per data source with extracted records."""
    display = source_name.replace("_", " ").title()
    ws = wb.create_sheet(display[:31])  # Sheet names max 31 chars
    ws.sheet_view.showGridLines = False

    ws.cell(2, 2, display).font = Font(name="Calibri", size=14, bold=True, color=NAVY)

    row = 4
    for extraction_type, records in data.items():
        if not records:
            continue

        _section_header(ws, extraction_type.replace("_", " ").upper(), row, 6)
        row += 1

        if not isinstance(records, list) or not records:
            ws.cell(row, 2, "No data extracted").font = _body_font(color=GREY)
            row += 2
            continue

        # Write headers from first record keys
        first = records[0] if isinstance(records[0], dict) else {}
        keys = list(first.keys())[:8]  # max 8 columns
        _set_col_widths(ws, [3] + [max(15, min(30, len(k) + 5)) for k in keys])

        for col, key in enumerate(keys, 2):
            c = ws.cell(row, col, key.replace("_", " ").title())
            c.font      = _header_font(size=9, color=WHITE)
            c.fill      = _fill(BLUE)
            c.alignment = _center()
            c.border    = _thin_border()
        ws.row_dimensions[row].height = 20
        row += 1

        for record in records[:200]:  # cap at 200 rows per extraction type
            if isinstance(record, dict):
                for col, key in enumerate(keys, 2):
                    val = record.get(key, "")
                    c  = ws.cell(row, col, val)
                    c.font      = _body_font()
                    c.alignment = _left()
                    c.border    = _thin_border()
            row += 1

        row += 1  # gap between extraction types


def _build_compliance(wb: Workbook, compliance: dict) -> None:
    ws = wb.create_sheet("Compliance")
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws, [3, 25, 55, 15])

    ws.cell(2, 2, "COMPLIANCE REVIEW").font = Font(name="Calibri", size=16, bold=True, color=NAVY)

    status = compliance.get("status", "unknown")
    status_color = {"passed": GREEN, "warnings": AMBER, "failed": RED}.get(status, GREY)
    ws.merge_cells("B4:D4")
    c = ws.cell(4, 2, f"Status: {status.upper()}")
    c.font      = Font(name="Calibri", size=12, bold=True, color=WHITE)
    c.fill      = _fill(status_color)
    c.alignment = _center()
    ws.row_dimensions[4].height = 28

    row = 6
    summary = [
        ("Checks Passed",  compliance.get("checks_passed", 0)),
        ("Checks Warned",  compliance.get("checks_warned", 0)),
        ("Checks Failed",  compliance.get("checks_failed", 0)),
        ("PII Hits",       len(compliance.get("pii_hits", []))),
    ]
    for label, value in summary:
        ws.cell(row, 2, label).font = _body_font(bold=True)
        ws.cell(row, 3, value).font = _body_font()
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1
    issues = compliance.get("issues", []) + compliance.get("warnings", [])
    if issues:
        _section_header(ws, "ISSUES & WARNINGS", row, 3)
        row += 1
        for issue in issues:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            c = ws.cell(row, 2, f"• {issue}")
            c.font      = _body_font()
            c.alignment = _left()
            ws.row_dimensions[row].height = 28
            row += 1


# ── Public API ─────────────────────────────────────────────────────────────────

def generate_excel(
    deliverables: dict,
    findings: list[dict],
    source_data: dict,
    target_company: str,
    pe_firm: str,
) -> bytes:
    """
    Generate a complete Excel workbook from KEEN pipeline output.

    Args:
        deliverables:   Dict from delivery agent state (executive_summary, detailed_report, etc.)
        findings:       List of finding dicts from all agents.
        source_data:    Raw extracted data keyed by source name.
        target_company: Name of the company being diligenced.
        pe_firm:        Name of the PE firm (client).

    Returns:
        Bytes of the .xlsx file.
    """
    wb = Workbook()

    # Cover
    _build_cover(wb, target_company, pe_firm, deliverables)

    # Executive summary
    exec_summary = deliverables.get("executive_summary", {})
    _build_executive_summary(wb, exec_summary)

    # All findings
    _build_findings(wb, findings)

    # One tab per data source
    for source_name, data in source_data.items():
        if data:
            try:
                _build_data_tab(wb, source_name, data)
            except Exception as exc:
                logger.warning("Excel: failed to build tab for '%s': %s", source_name, exc)

    # Compliance
    compliance = deliverables.get("compliance", {})
    if compliance:
        _build_compliance(wb, compliance)

    # Serialise to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    logger.info(
        "Excel: generated workbook with %d sheets for '%s'",
        len(wb.worksheets), target_company,
    )
    return buf.read()
